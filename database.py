import sqlite3
from datetime import datetime
import os

def recreate_database():
    """Пересоздает базу данных с нуля"""
    if os.path.exists('chillman.db'):
        os.remove('chillman.db')
    init_database()

def init_database():
    # Создаем соединение с базой данных
    conn = sqlite3.connect('chillman.db')
    cursor = conn.cursor()
    
    # Создаем таблицы
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'Работник',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Таблица с основной информацией о клиентах
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_info (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            middle_name TEXT,
            birth_date DATE,
            phone TEXT,
            email TEXT,
            address TEXT,
            last_call DATETIME,
            last_caller TEXT,
            last_offer TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Таблица с финансовой информацией клиентов
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS client_financial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL,
            tariff TEXT DEFAULT 'Стандарт',
            balance REAL DEFAULT 0.00,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (client_id) REFERENCES client_info(id) ON DELETE CASCADE
        )
    ''')
    
    # Таблица с акциями
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS promotions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            offer_text TEXT NOT NULL
        )
    ''')
    
    # Добавляем акции, если их еще нет
    cursor.execute('SELECT COUNT(*) FROM promotions')
    if cursor.fetchone()[0] == 0:
        promotions = [
            ("Скидка 20% на услуги на 3 месяца",),
            ("Кэшбек 2% навсегда",),
            ("Приведи друга - получи месяц бесплатно",),
            ("Бесплатный апгрейд тарифа на 6 месяцев",),
            ("Удвоенные баллы лояльности на год",)
        ]
        cursor.executemany('INSERT INTO promotions (offer_text) VALUES (?)', promotions)
    
    # Добавляем новые столбцы, если их нет
    try:
        cursor.execute('SELECT last_call FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_call DATETIME')
        
    try:
        cursor.execute('SELECT last_caller FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_caller TEXT')
        
    try:
        cursor.execute('SELECT last_offer FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_offer TEXT')
    
    # Проверяем, есть ли уже пользователь admin
    cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', ('admin',))
    if cursor.fetchone()[0] == 0:
        # Добавляем администратора по умолчанию
        cursor.execute('''
            INSERT INTO users (username, password, role)
            VALUES (?, ?, ?)
        ''', ('admin', 'admin123', 'admin'))
    
    # Миграция данных из старой таблицы, если она существует
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clients'")
    if cursor.fetchone() is not None:
        print("Migrating data from old table structure...")
        cursor.execute('''
            INSERT INTO client_info (
                id, last_name, first_name, middle_name, birth_date,
                phone, email, address, created_at, updated_at
            )
            SELECT 
                id, last_name, first_name, middle_name, birth_date,
                phone, email, address, created_at, updated_at
            FROM clients
        ''')
        
        cursor.execute('''
            INSERT INTO client_financial (client_id, tariff, balance)
            SELECT id, tariff, balance FROM clients
        ''')
        
        # Удаляем старую таблицу
        cursor.execute('DROP TABLE clients')
    
    conn.commit()
    conn.close()

def migrate_database():
    """Добавляет новые столбцы в существующую базу данных"""
    conn = sqlite3.connect('clients.db')
    cursor = conn.cursor()
    
    # Добавляем новые столбцы, если их нет
    try:
        cursor.execute('SELECT last_call FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_call DATETIME')
        print("Добавлен столбец last_call")
        
    try:
        cursor.execute('SELECT last_caller FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_caller TEXT')
        print("Добавлен столбец last_caller")
        
    try:
        cursor.execute('SELECT last_offer FROM client_info LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE client_info ADD COLUMN last_offer TEXT')
        print("Добавлен столбец last_offer")
    
    conn.commit()
    conn.close()

def migrate_from_excel():
    """
    Функция для миграции данных из Excel в SQLite
    """
    try:
        from openpyxl import load_workbook
        import sqlite3
        
        # Подключаемся к базе данных
        conn = sqlite3.connect('chillman.db')
        cursor = conn.cursor()
        
        # Миграция пользователей из file3.xlsx
        if os.path.exists('file3.xlsx'):
            print("Начинаем миграцию пользователей из file3.xlsx...")
            wb = load_workbook('file3.xlsx')
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if row and len(row) >= 3 and row[0] and row[1] and row[2]:  # Проверяем наличие всех необходимых данных
                        username = str(row[1]).strip()
                        password = str(row[2]).strip()
                        role = str(row[3]).strip() if len(row) > 3 and row[3] else 'Работник'
                        
                        try:
                            cursor.execute('''
                                INSERT INTO users (username, password, role)
                                VALUES (?, ?, ?)
                            ''', (username, password, 'admin' if role.lower() == 'admin' else 'Работник'))
                            print(f"Добавлен пользователь: {username}")
                        except sqlite3.IntegrityError:
                            print(f"Пользователь {username} уже существует")
                except Exception as e:
                    print(f"Ошибка при обработке строки пользователя: {e}")
                    continue
        
        # Миграция клиентов из file2.xlsx
        if os.path.exists('file2.xlsx'):
            print("\nНачинаем миграцию клиентов из file2.xlsx...")
            wb = load_workbook('file2.xlsx')
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if row and len(row) >= 2 and row[1]:  # Проверяем наличие хотя бы имени клиента
                        name = str(row[1]).strip()
                        phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                        email = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        address = str(row[4]).strip() if len(row) > 4 and row[4] else None
                        
                        try:
                            cursor.execute('''
                                INSERT INTO client_info (last_name, first_name, phone, email, address)
                                VALUES (?, ?, ?, ?, ?)
                            ''', (name, '', phone, email, address))
                            print(f"Добавлен клиент: {name}")
                        except sqlite3.IntegrityError:
                            print(f"Клиент {name} уже существует")
                except Exception as e:
                    print(f"Ошибка при обработке строки клиента: {e}")
                    continue
        
        conn.commit()
        conn.close()
        print("\nМиграция данных завершена успешно!")
        return True
        
    except Exception as e:
        print(f"Ошибка при миграции данных: {e}")
        return False

if __name__ == "__main__":
    init_database()
    migrate_database()  # Добавляем вызов миграции
    migrate_from_excel()
